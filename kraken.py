import os
import re
import cv2
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from kraken import binarization, pageseg, rpred
from PIL import Image

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["OUTPUT_FOLDER"] = OUTPUT_FOLDER

# Define regex patterns
main_pattern = r"\b\d{2}-[A-Z]-[A-Z]{3}\d-[A-Z]{1,3}-[A-Z]{2}\d+\b"
tag_pattern = r"\b\d{2}-[A-Z]-[A-Z]{3}\d-([A-Z]{1,3})-[A-Z]{2}\d+\b"

# Load the master tag list from the provided file
MASTER_FILE_PATH = "/Users/harshvith/Downloads/TAG IDENTIFIER CODES-MASTER.xlsx"

if os.path.exists(MASTER_FILE_PATH):
    master_df = pd.read_excel(MASTER_FILE_PATH, sheet_name="TAG IDENTIFIER")

    # Ensure case consistency, trim spaces, and drop duplicates
    master_df["TAG IDENTIFIER CODE"] = master_df["TAG IDENTIFIER CODE"].astype(str).str.strip().str.upper()
    master_df = master_df.drop_duplicates(subset=["TAG IDENTIFIER CODE"], keep="first")

    # Convert to dictionary
    tag_mapping = master_df.set_index("TAG IDENTIFIER CODE")[["PRIMARY EQUIPMENT DESCRIPTION", "DEPARTMENT"]].to_dict("index")

    print(f"Loaded Master Tags: {len(tag_mapping)} unique entries")
else:
    print("Master file not found!")
    tag_mapping = {}

from kraken.lib.models import load_any

def kraken_ocr(image):
    """Uses Kraken OCR to extract text from an image."""
    model_path = "best"  # Ensure model is correctly loaded
    model = load_any(model_path)  # Load the actual model instead of using a string
    
    img_pil = Image.fromarray(image)

    # Convert image to binarized format for better OCR accuracy
    bw_img = binarization.nlbin(img_pil)
    segmentation = pageseg.segment(bw_img)

    # Perform OCR using Kraken's correct function call
    results = rpred.rpred(model, bw_img, segmentation)

    extracted_text = " ".join([r.prediction for r in results if r.prediction])
    return extracted_text




def process_pdf(pdf_path, output_excel_path):
    """Processes the uploaded PDF to extract text and unique tags, saving them in an Excel file."""
    
    drawing_number = os.path.splitext(os.path.basename(pdf_path))[0]

    images = convert_from_path(pdf_path, dpi=300, thread_count=4)
    extracted_data = []
    extracted_tags = []

    for page_number, img in enumerate(images, start=1):
        image_path = os.path.join(UPLOAD_FOLDER, f"page_{page_number}.png")
        img.save(image_path, "PNG")
        
        image = cv2.imread(image_path)
        hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

        # Define red color range for text detection
        lower_red1, upper_red1 = np.array([0, 100, 100]), np.array([10, 255, 255])
        lower_red2, upper_red2 = np.array([170, 100, 100]), np.array([180, 255, 255])
        
        # Create masks for detecting red text
        mask = cv2.inRange(hsv, lower_red1, upper_red1) + cv2.inRange(hsv, lower_red2, upper_red2)
        red_text_only = cv2.bitwise_and(image, image, mask=mask)

        # Convert to grayscale and apply thresholding
        gray = cv2.cvtColor(red_text_only, cv2.COLOR_BGR2GRAY)
        gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
        
        # Perform OCR using Kraken
        extracted_text = kraken_ocr(gray)

        # Process extracted text
        words = extracted_text.split()
        for word in words:
            if re.fullmatch(main_pattern, word):
                tag_match = re.match(tag_pattern, word)
                unique_tag = tag_match.group(1) if tag_match else None

                if unique_tag:
                    unique_tag = unique_tag.strip().upper()

                    if unique_tag in tag_mapping:
                        equipment_desc = tag_mapping[unique_tag]["PRIMARY EQUIPMENT DESCRIPTION"]
                        department = tag_mapping[unique_tag]["DEPARTMENT"]
                    else:
                        equipment_desc = "N/A"
                        department = "N/A"
                else:
                    equipment_desc = "N/A"
                    department = "N/A"

                extracted_data.append({
                    "Sl.no.": len(extracted_data) + 1,
                    "Discipline": department,
                    "Tag Number": word, 
                    "Tag Identifier Code": unique_tag,
                    "Equipment Description": equipment_desc,
                    "Drawing Number": drawing_number,
                    "Sheet No.": page_number
                })

                if unique_tag:
                    extracted_tags.append(unique_tag)

    df_text = pd.DataFrame(extracted_data)
    df_tags = pd.DataFrame(extracted_tags, columns=["Tag"])
    tag_counts = df_tags["Tag"].value_counts().reset_index()
    tag_counts.columns = ["Tag", "Count"]
    tag_counts.insert(0, "S.No.", range(1, len(tag_counts) + 1))

    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df_text.to_excel(writer, sheet_name="Extracted Text", index=False)
        tag_counts.to_excel(writer, sheet_name="Tag Counts", index=False)

        # Load the workbook and apply center alignment
    wb = load_workbook(output_excel_path)

    # Adjust "Extracted Text" sheet
    if "Extracted Text" in wb.sheetnames:
        ws_text = wb["Extracted Text"]
        for col in ws_text.columns:
            max_length = max((len(str(cell.value)) for cell in col), default=0)
            ws_text.column_dimensions[col[0].column_letter].width = max_length + 2

    # Adjust "Tag Counts" sheet
    if "Tag Counts" in wb.sheetnames:
        ws_tags = wb["Tag Counts"]
        for col in ws_tags.columns:
            max_length = max((len(str(cell.value)) for cell in col), default=0)
            ws_tags.column_dimensions[col[0].column_letter].width = max_length + 2

    def center_align(sheet_name):
        """Applies center alignment to all cells in a given sheet."""
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply center alignment to both sheets
    center_align("Extracted Text")
    center_align("Tag Counts")

    # Save the workbook with formatted cells
    wb.save(output_excel_path)

@app.route("/upload", methods=["POST"])
def upload_file():
    file = request.files["file"]
    file_path = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
    file.save(file_path)

    output_excel_path = os.path.join(app.config["OUTPUT_FOLDER"], "extracted_data.xlsx")
    process_pdf(file_path, output_excel_path)

    return jsonify({"message": "File processed successfully", "download_url": "/download"}), 200

@app.route("/download", methods=["GET"])
def download_file():
    output_excel_path = os.path.join(app.config["OUTPUT_FOLDER"], "extracted_data.xlsx")
    
    if not os.path.exists(output_excel_path):
        return jsonify({"error": "File not found"}), 404

    return send_file(output_excel_path, as_attachment=True)


@app.route("/tags", methods=["GET"])
def get_tags():
    """Returns extracted unique tags and their counts."""
    output_excel_path = os.path.join(app.config["OUTPUT_FOLDER"], "extracted_data.xlsx")
    
    if not os.path.exists(output_excel_path):
        return jsonify({"tags": []})

    df = pd.read_excel(output_excel_path, sheet_name="Tag Counts")

    if df.empty:
        return jsonify({"tags": []})

    tags = df.to_dict(orient="records")
    return jsonify({"tags": tags})

if __name__ == "__main__":
    app.run(debug=True)
